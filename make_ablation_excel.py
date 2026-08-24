#!/usr/bin/env python3
"""ablation 결과(로그+CSV) → 점수 항별 성능 변화 Excel.
사용: python make_ablation_excel.py <abl.log> <ablation_by_group.csv> <out.xlsx>"""
import re, sys, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

LOG, GCSV, OUT = sys.argv[1], sys.argv[2], sys.argv[3]

HD = Font(bold=True, color="FFFFFF"); HF = PatternFill("solid", fgColor="305496")
YL = PatternFill("solid", fgColor="FFF2CC"); GD = PatternFill("solid", fgColor="C6EFCE")
RD = PatternFill("solid", fgColor="FFC7CE"); BB = Border(*[Side(style="thin", color="D9D9D9")] * 4)
C = Alignment(horizontal="center"); WL = Alignment(vertical="center", wrap_text=True)


def hdr(ws, r=1):
    for c in ws[r]:
        c.font = HD; c.fill = HF; c.alignment = C; c.border = BB


# ── 로그 파싱: 설정별 정확도 ──
cfg = {}
for ln in open(LOG, encoding="utf-8").read().splitlines():
    m = re.match(r"\s{2}(\S.*?)\s{2,}([\d.]+)%\s+\((\d+)/(\d+)\)", ln)
    if m:
        cfg[m.group(1).strip()] = (float(m.group(2)), int(m.group(3)), int(m.group(4)))
v3 = cfg.get("v3  (시각항 OFF, 기준선)", (0, 0, 1))[0]
v4 = cfg.get("v4  (전부 ON)", (0, 0, 1))[0]

wb = Workbook()

# ── Sheet1: 설정별 정확도 ──
ws = wb.active; ws.title = "설정별 정확도"
ws.append(["설정", "정확도(%)", "v3 대비(%p)", "맞은/전체"]); hdr(ws)
for name, (a, c, t) in cfg.items():
    ws.append([name, a, round(a - v3, 1), f"{c}/{t}"])
    row = ws[ws.max_row]
    if name.startswith("v3  ") or name.startswith("v4  "):
        for cc in row:
            cc.fill = YL; cc.font = Font(bold=True)
for col, w in zip("ABCD", [26, 12, 12, 14]):
    ws.column_dimensions[col].width = w
for r in ws.iter_rows(min_row=2):
    for cc in r:
        cc.border = BB; cc.alignment = C

# ── Sheet2: 항별 기여 (단독/필요성) ──
ws2 = wb.create_sheet("항별 기여")
ws2.append(["점수 항", "분류", "단독 기여(v3+항 − v3)", "필요성(v4−항 하락폭)"]); hdr(ws2)

def g(name):
    return cfg.get(name, (None,))[0]

rows = [
    ("빈도수",         "핵심(v3)", None, (v4 - g("v4 − 빈도수")) if g("v4 − 빈도수") is not None else None),
    ("값다양성",       "핵심(v3)", None, (v4 - g("v4 − 값다양성")) if g("v4 − 값다양성") is not None else None),
    ("크기변동penalty", "핵심(v3)", None, (v4 - g("v4 − 크기변동penalty")) if g("v4 − 크기변동penalty") is not None else None),
    ("2곳일치",        "핵심(v3)", None, (v4 - g("v4 − 2곳일치")) if g("v4 − 2곳일치") is not None else None),
    ("크기(시각)",     "시각(v4)", (g("v3 + 크기") - v3) if g("v3 + 크기") is not None else None,
     (v4 - g("v4 − 크기")) if g("v4 − 크기") is not None else None),
    ("배경하이라이트",   "시각(v4)", (g("v3 + 배경하이라이트") - v3) if g("v3 + 배경하이라이트") is not None else None,
     (v4 - g("v4 − 배경하이라이트")) if g("v4 − 배경하이라이트") is not None else None),
    ("텍스트대비",     "시각(v4)", (g("v3 + 대비") - v3) if g("v3 + 대비") is not None else None,
     (v4 - g("v4 − 대비")) if g("v4 − 대비") is not None else None),
    ("전역현저성",     "시각(v4)", (g("v3 + 전역현저성") - v3) if g("v3 + 전역현저성") is not None else None,
     (v4 - g("v4 − 전역현저성")) if g("v4 − 전역현저성") is not None else None),
    ("채도",          "시각(v4)", (g("v3 + 채도") - v3) if g("v3 + 채도") is not None else None, None),
]
for name, kind, solo, need in rows:
    ws2.append([name, kind, None if solo is None else round(solo, 1),
                None if need is None else round(need, 1)])
    r = ws2[ws2.max_row]
    hi = max([x for x in (solo, need) if x is not None] or [0])
    if hi >= 3:
        for cc in r:
            cc.fill = GD
    elif hi <= 0.05 and (solo is not None or need is not None):
        for cc in r:
            cc.fill = RD
for col, w in zip("ABCD", [16, 12, 22, 22]):
    ws2.column_dimensions[col].width = w
for r in ws2.iter_rows(min_row=2):
    for cc in r:
        cc.border = BB; cc.alignment = C

# ── Sheet3: 사용자 공식 매핑 ──
ws3 = wb.create_sheet("공식 항 매핑")
ws3.append(["Score ∝ 항", "코드에 있나", "ablation 됨", "측정 결과"]); hdr(ws3)
form = [
    ("빈도수", "O (v3 핵심)", "O", f"빼면 {round(v4-g('v4 − 빈도수'),1) if g('v4 − 빈도수') is not None else '-'}%p 하락"),
    ("값다양성", "O (v3 핵심)", "O", f"빼면 {round(v4-g('v4 − 값다양성'),1) if g('v4 − 값다양성') is not None else '-'}%p 하락"),
    ("1/위치변동성", "X (점수항 없음)", "X", "위치는 '묶는 기준'이지 점수항 아님"),
    ("1/크기변동성", "O (v3 핵심)", "O", f"빼면 {round(v4-g('v4 − 크기변동penalty'),1) if g('v4 − 크기변동penalty') is not None else '-'}%p 하락"),
    ("텍스트 대비", "O (v4)", "O", f"단독 {round(g('v3 + 대비')-v3,1) if g('v3 + 대비') is not None else '-'}%p / 빼면 {round(v4-g('v4 − 대비'),1) if g('v4 − 대비') is not None else '-'}%p"),
]
for a, b, c, d in form:
    ws3.append([a, b, c, d])
    if "X" in c:
        for cc in ws3[ws3.max_row]:
            cc.fill = RD
for col, w in zip("ABCD", [16, 18, 12, 40]):
    ws3.column_dimensions[col].width = w
for r in ws3.iter_rows(min_row=2):
    for cc in r:
        cc.border = BB; cc.alignment = WL

# ── Sheet4: UI별 표 ──
ws4 = wb.create_sheet("UI별")
try:
    rows = list(csv.reader(open(GCSV, encoding="utf-8-sig")))
    for i, r in enumerate(rows):
        ws4.append(r)
    hdr(ws4)
    for col in range(1, len(rows[0]) + 1 if rows else 1):
        ws4.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 10
except Exception as e:
    ws4.append([f"CSV 읽기 실패: {e}"])

wb.save(OUT)
print("saved", OUT)
