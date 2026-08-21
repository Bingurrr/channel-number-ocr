#!/usr/bin/env python3
"""bench_v6 결과 → v4/v5/v6 비교 Excel."""
import json, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SC = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("bench_v6_full")
OUT = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("v4_v5_v6_결과.xlsx")

so = json.load(open(SC / "summary_horig.json"))
s2 = json.load(open(SC / "summary_h270.json"))

wb = Workbook()
HD = Font(bold=True, color="FFFFFF"); HF = PatternFill("solid", fgColor="305496")
GD = PatternFill("solid", fgColor="C6EFCE"); RD = PatternFill("solid", fgColor="FFC7CE")
YL = PatternFill("solid", fgColor="FFF2CC")
BB = Border(*[Side(style="thin", color="D9D9D9")] * 4)
C = Alignment(horizontal="center"); WL = Alignment(vertical="center", wrap_text=True)


def hdr(ws, row=1):
    for c in ws[row]:
        c.font = HD; c.fill = HF; c.alignment = C; c.border = BB


# ── 종합 ──
ws = wb.active; ws.title = "종합"
ws.append(["항목", "v4", "v5", "v6"]); hdr(ws)
ret = lambda d, t: round(s2[t]['e2e_acc'] / max(1e-9, so[t]['e2e_acc']) * 100, 1)
rows = [
    ["구조", "매프레임 full OCR + json", "5프레임후 rec-only(락 고정)", "v5 + 신뢰도게이트 + 주기재검증"],
    ["── 원본 720 ──", "", "", ""],
    ["정확도(E2E)", f"{so['v4']['e2e_acc']}%", f"{so['v5']['e2e_acc']}%", f"{so['v6']['e2e_acc']}%"],
    ["읽음율", f"{so['v4']['read_rate']}%", f"{so['v5']['read_rate']}%", f"{so['v6']['read_rate']}%"],
    ["처리시간", f"{so['v4']['total_time_s']}s", f"{so['v5']['total_time_s']}s", f"{so['v6']['total_time_s']}s"],
    ["속도", "1.0x", f"{so['speedup_v5']}x", f"{so['speedup_v6']}x"],
    ["full OCR 호출", f"{so['v4']['frames']}", f"{so['folders']*so['window']}", f"{so['v6_full_ocr_calls']}"],
    ["── 저해상 270 ──", "", "", ""],
    ["정확도(E2E)", f"{s2['v4']['e2e_acc']}%", f"{s2['v5']['e2e_acc']}%", f"{s2['v6']['e2e_acc']}%"],
    ["처리시간", f"{s2['v4']['total_time_s']}s", f"{s2['v5']['total_time_s']}s", f"{s2['v6']['total_time_s']}s"],
    ["속도", "1.0x", f"{s2['speedup_v5']}x", f"{s2['speedup_v6']}x"],
    ["720→270 유지율", f"{ret(s2,'v4')}%", f"{ret(s2,'v5')}%", f"{ret(s2,'v6')}%"],
]
for r in rows:
    ws.append(r)
    if r[0].startswith("──"):
        for c in ws[ws.max_row]:
            c.font = Font(bold=True); c.fill = YL
for col, w in zip("ABCD", [16, 24, 26, 30]):
    ws.column_dimensions[col].width = w
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.border = BB; c.alignment = WL

# ── UI별 720 ──
ws2 = wb.create_sheet("UI별_720")
ws2.append(["UI", "프레임", "v4%", "v5%", "v6%", "v6-v4", "v6-v5", "v6 fullOCR", "자가교정", "v4 t", "v6 t"])
hdr(ws2)
for r in so["per_ui"]:
    d64 = round(r["v6"] - r["v4"], 1); d65 = round(r["v6"] - r["v5"], 1)
    ws2.append([r["ui"], r["n"], r["v4"], r["v5"], r["v6"], d64, d65,
                r["v6_full"], r["v6_heal"], r["v4_t"], r["v6_t"]])
    if d65 >= 10:
        for c in ws2[ws2.max_row]:
            c.fill = GD                      # v6가 v5 크게 복구
    elif d64 <= -5:
        for c in ws2[ws2.max_row]:
            c.fill = RD                      # v6가 v4보다 낮음
for col, w in zip("ABCDEFGHIJK", [8, 8, 7, 7, 7, 8, 8, 11, 9, 8, 8]):
    ws2.column_dimensions[col].width = w
for row in ws2.iter_rows(min_row=2):
    for c in row:
        c.border = BB; c.alignment = C

# ── 방법론 ──
ws3 = wb.create_sheet("방법론")
notes = [
    ["v6 = v5 + 2가지", ""],
    ["① 락 신뢰도 게이트", "phase A 클러스터 coverage가 conf_min(0.5) 미만이면 rec-only로 안 넘어가고 full OCR 유지(=v4 폴백). 애매한 폴더에서 회귀 방지"],
    ["② 주기적 재검증", "rec-only 모드에서도 recheck_every(20)프레임마다 full OCR 1장을 섞어 누적버퍼에 추가→재클러스터. 값 다양성이 쌓이며 초반 오락 ROI를 자가교정"],
    ["③ 저신뢰 읽기 트리거", "rec 점수 < min_score(0.30)면 그 프레임 즉시 full OCR 재검증"],
    ["비용", f"full OCR = window(5) + 100/recheck ≈ 폴더당 {so['v6_full_ocr_calls']//max(1,so['folders'])}장 (v4는 100장). rec-only가 대부분 → 여전히 크게 빠름"],
    ["", ""],
    ["결과 요약", ""],
    ["v5 붕괴 복구", "UI_08(0→98 등) 초반 오락 폴더를 주기재검증이 되살림"],
    ["v6 vs v4", f"정확도 {so['v6']['e2e_acc']}% vs {so['v4']['e2e_acc']}% 이면서 {so['speedup_v6']}x 빠름"],
    ["남는 실패", "UI_35/40은 v4/v5/v6 모두 낮음 = 흰-on-주황 하이라이트 '읽기'실패(선택 아닌 rec 문제, 파인튜닝 대상)"],
    ["저해상 270", f"v6 유지율 {ret(s2,'v6')}% — 해상도 낮춰도 강건"],
]
ws3.append(["항목", "설명"]); hdr(ws3)
for r in notes:
    ws3.append(r)
    if r[1] == "":
        for c in ws3[ws3.max_row]:
            c.font = Font(bold=True); c.fill = YL
ws3.column_dimensions["A"].width = 18; ws3.column_dimensions["B"].width = 88
for row in ws3.iter_rows(min_row=2):
    for c in row:
        c.border = BB; c.alignment = WL

wb.save(OUT)
print("saved", OUT)
