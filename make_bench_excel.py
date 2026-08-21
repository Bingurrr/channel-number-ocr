#!/usr/bin/env python3
"""bench_v5 결과(summary json) → v4 vs v5 비교 Excel."""
import json, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SC = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("bench_full")
OUT = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("v4_vs_v5_결과.xlsx")

so = json.load(open(SC / "summary_horig.json"))
s2 = json.load(open(SC / "summary_h270.json"))

wb = Workbook()
HD = Font(bold=True, color="FFFFFF"); HF = PatternFill("solid", fgColor="305496")
GD = PatternFill("solid", fgColor="E2EFDA"); RD = PatternFill("solid", fgColor="FCE4D6")
BB = Border(*[Side(style="thin", color="D9D9D9")] * 4)
C = Alignment(horizontal="center")


def style_header(ws, row=1):
    for c in ws[row]:
        c.font = HD; c.fill = HF; c.alignment = C; c.border = BB


# ── Sheet1: 종합 ──
ws = wb.active; ws.title = "종합"
ws.append(["항목", "v4 (매프레임 full OCR)", "v5 (window후 rec-only)"])
style_header(ws)
rows = [
    ["구조", "det+rec 매 프레임 + json 경유", "앞 5프레임만 full OCR→ROI락, 이후 rec-only, in-memory"],
    ["── 원본 720 ──", "", ""],
    ["정확도(E2E)", f"{so['v4']['e2e_acc']}%", f"{so['v5']['e2e_acc']}%"],
    ["읽음율(read)", f"{so['v4']['read_rate']}%", f"{so['v5']['read_rate']}%"],
    ["총 처리시간", f"{so['v4']['total_time_s']}s", f"{so['v5']['total_time_s']}s"],
    ["속도", "1.0x (기준)", f"{so['speedup']}x 빠름"],
    ["full OCR 호출", f"{so['full_ocr_calls']}회", f"{so['window']}/폴더 = {so['folders']*so['window']}회"],
    ["rec-only 호출", "0", f"{so['rec_only_calls']}회"],
    ["── 저해상 270 ──", "", ""],
    ["정확도(E2E)", f"{s2['v4']['e2e_acc']}%", f"{s2['v5']['e2e_acc']}%"],
    ["읽음율(read)", f"{s2['v4']['read_rate']}%", f"{s2['v5']['read_rate']}%"],
    ["총 처리시간", f"{s2['v4']['total_time_s']}s", f"{s2['v5']['total_time_s']}s"],
    ["속도", "1.0x (기준)", f"{s2['speedup']}x 빠름"],
    ["── 720→270 유지율 ──",
     f"{round(s2['v4']['e2e_acc']/max(1e-9,so['v4']['e2e_acc'])*100,1)}%",
     f"{round(s2['v5']['e2e_acc']/max(1e-9,so['v5']['e2e_acc'])*100,1)}%"],
]
for r in rows:
    ws.append(r)
    if r[0].startswith("──"):
        for c in ws[ws.max_row]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFF2CC")
for col, w in zip("ABC", [20, 34, 52]):
    ws.column_dimensions[col].width = w
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.border = BB; c.alignment = Alignment(vertical="center", wrap_text=True)

# ── Sheet2: UI별 (720) ──
ws2 = wb.create_sheet("UI별_720")
ws2.append(["UI", "프레임", "v4 정확도%", "v5 정확도%", "차이(v5-v4)", "ROI락", "v4 시간s", "v5 시간s", "판정"])
style_header(ws2)
for r in so["per_ui"]:
    diff = round(r["v5_acc"] - r["v4_acc"], 1)
    if r["v4_acc"] == 0 and r["v5_acc"] == 0:
        verdict = "둘다실패(읽기)"
    elif diff <= -10:
        verdict = "v5 회귀(ROI오락)"
    elif diff >= 10:
        verdict = "v5 개선(강제읽기)"
    else:
        verdict = "동등"
    ws2.append([r["ui"], r["n"], r["v4_acc"], r["v5_acc"], diff,
                "O" if r["box"] else "X", r["v4_t"], r["v5_t"], verdict])
    fill = GD if diff >= 10 else (RD if diff <= -10 else None)
    if fill:
        for c in ws2[ws2.max_row]:
            c.fill = fill
for col, w in zip("ABCDEFGHI", [8, 8, 11, 11, 12, 7, 9, 9, 18]):
    ws2.column_dimensions[col].width = w
for row in ws2.iter_rows(min_row=2):
    for c in row:
        c.border = BB; c.alignment = C

# ── Sheet3: 방법론 ──
ws3 = wb.create_sheet("방법론")
notes = [
    ["v5 알고리즘", ""],
    ["1) phase A", "폴더(시퀀스) 앞 window(5)프레임은 v4처럼 full OCR(det+rec) 진행"],
    ["2) 클러스터 락", "5프레임 누적으로 slot_v4가 채널 '위치(ROI)'를 확정 (값이 아니라 위치)"],
    ["3) phase B", "이후 모든 프레임은 det 생략, 그 ROI만 crop→확대→rec-only로 직접 읽음"],
    ["왜 빠른가", "det(무거움)를 폴더당 5회만. 나머지는 작은 ROI에 rec만 → 6x↑"],
    ["왜 위치 락이 되나", "Test_overlay는 폴더 내 UI 디자인 고정=채널 위치 고정, 값만 프레임마다 변주"],
    ["json 경유 제거", "v4 드라이버는 full_ocr.json 저장/로드(연구용 배치). v5는 in-memory 스트리밍=온디바이스형"],
    ["", ""],
    ["v5 리스크", ""],
    ["ROI 오락", "앞 5프레임 클러스터가 틀리면 이후 전부 엉뚱한 곳 읽음(UI_08/21/39). v4는 매프레임 재평가라 없음"],
    ["대응책", "락 후 주기적 재검증 / 다중 프레임 합의 상승 / 락 신뢰도 게이트로 실패 시 full OCR 폴백"],
    ["둘다 실패 폴더", "UI_35/UI_40은 v4도 0% = 흰-on-주황 하이라이트 읽기실패(방법 무관, rec 파인튜닝 이슈)"],
]
ws3.append(["항목", "설명"]); style_header(ws3)
for r in notes:
    ws3.append(r)
    if r[1] == "":
        for c in ws3[ws3.max_row]:
            c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="FFF2CC")
ws3.column_dimensions["A"].width = 16; ws3.column_dimensions["B"].width = 82
for row in ws3.iter_rows(min_row=2):
    for c in row:
        c.border = BB; c.alignment = Alignment(vertical="center", wrap_text=True)

wb.save(OUT)
print("saved", OUT)
